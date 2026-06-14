#!/usr/bin/env python3
"""Import byte-matched C functions from the prior C:\\NFS4 decomp attempt.

The prior attempt (C:\\NFS4) byte-matched 173 functions as real C (Kind=C in
nfs4-compile.xlsx) against the identical binary, using the identical toolchain
(ccpsx -Xm -O2 -G4). Those .c files are self-contained (no includes). We splice
each into our src/main.c, replacing its INCLUDE_ASM line with the real body.

Kind=asm "matches" (.word literal shims) are intentionally NOT imported — they
are disguised asm, not honest decompilation.

Run, then rebuild + objdiff to confirm the matched count rose.
"""
import openpyxl
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
PRIOR = Path(r"C:\NFS4\decomp")
XLSX = Path(r"C:\NFS4\nfs4-compile.xlsx")
MAIN = ROOT / "src" / "main.c"


def kind_c_matches():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["functions"]
    hdr = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[ix["Status"]] == "MATCH" and r[ix["Kind"]] == "C":
            out.append((r[ix["Symbol"]], r[ix["File"]]))
    return out


def clean_body(text):
    """Strip leading /* ... */ banner comment(s); keep the definition."""
    # remove block comments
    text = re.sub(r"/\*.*?\*/", "", text, flags=re.S)
    return text.strip()


def main():
    src = MAIN.read_text()
    imported = skipped_nostub = skipped_asm = 0
    for sym, fn in kind_c_matches():
        p = PRIOR / fn
        if not p.exists():
            continue
        body = clean_body(p.read_text(errors="replace"))
        if "__asm__" in body or ".word" in body:
            skipped_asm += 1
            continue
        # match this symbol's INCLUDE_ASM line in our monolithic main.c
        pat = re.compile(
            r'^INCLUDE_ASM\("asm/nonmatchings/main", ' + re.escape(sym) + r'\);\s*$',
            re.M)
        if not pat.search(src):
            skipped_nostub += 1  # already a C stub, or name differs
            continue
        src = pat.sub(lambda m: body, src, count=1)
        imported += 1
    MAIN.write_text(src)
    print(f"imported {imported} bodies; skipped {skipped_nostub} (no INCLUDE_ASM "
          f"line / already C), {skipped_asm} (asm-shim)")


if __name__ == "__main__":
    main()
